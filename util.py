import customtkinter as ctk
from dados.tabelas import User, Conta, session, Conta_dados
from CTkMessagebox import CTkMessagebox
from PIL import Image
from customtkinter import CTkImage
import pandas as pd
from tkinter import filedialog
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def limpar_tela(frame_atual):
    
    for widget in frame_atual.winfo_children():
        widget.destroy()

def limpar_area_principal(area_principal):
    for widget in area_principal.winfo_children():
        widget.destroy()

def coletar_dados():
    contas = session.query(Conta).all()

    dados = []

    for dado in contas:
        dados.append({
            'Nome': dado.nome,
            'Email': dado.email,
            'Ultimo Backup': dado.ultimo_bkp,
            'Segundo Backup': dado.segundo_backup if dado.segundo_backup else '',
            'Cor': dado.cor_ultimo_backup
        })
    
    return pd.DataFrame(dados)


def exportar_para_excel(df, janela):
    cores_status = {
        'green': 'FF008000',
        'yellow': 'FFFFFF00',
        'red': 'FFFF0000'
    }

    borda_preta = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    caminho = filedialog.asksaveasfilename(parent=janela, defaultextension=".xlsx",
                                          filetypes=[("Arquivos Excel", "*.xlsx")],
                                          title="Salvar como")

    if caminho:
        with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
            df.drop(columns=['Cor'], inplace=False).to_excel(writer, index=False, sheet_name='Backup')  # Não exporta coluna 'Cor'
            planilha = writer.sheets['Backup']

            tamanhos_colunas = {
                'Nome': 40,
                'Email': 45,
                'Ultimo Backup': 35,
                'Segundo Backup': 35
            }

            for idx, col in enumerate(df.drop(columns=['Cor']).columns, 1):
                col_letter = get_column_letter(idx)
                largura = tamanhos_colunas.get(col, 20)
                planilha.column_dimensions[col_letter].width = largura

            num_linhas = df.shape[0] + 1  
            num_colunas = df.shape[1] - 1  # menos a coluna 'Cor'

            for row in range(1, num_linhas + 1):
                for col in range(1, num_colunas + 1):
                    celula = planilha.cell(row=row, column=col)
                    celula.border = borda_preta

            col_index = list(df.columns).index("Ultimo Backup") + 1

            for i, cor_valor in enumerate(df['Cor'], start=2):  # começa da linha 2
                cor_hex = cores_status.get(cor_valor, 'FFFFFFFF')  # branco se não tiver cor
                celula = planilha.cell(row=i, column=col_index)
                fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type='solid')
                celula.fill = fill
                celula.border = borda_preta


def b_exportar(janela):
    df = coletar_dados()
    exportar_para_excel(df, janela)


def coletar_dados2():
    contas = session.query(Conta_dados).all()

    dados = []

    for dado in contas:
        dados.append({
            'Nome': dado.nome,
            'Email': dado.email,
            'Ultimo Backup': dado.ultimo_bkp,
            'OBS': dado.obs,
            'Cor': dado.cor_ultimo_backup
        })
    
    return pd.DataFrame(dados)


def b_exportar2(janela):
    df = coletar_dados2()
    exportar_para_excel(df, janela)

def importar_ultima_alteracao():
    # Abre o seletor de arquivos para escolher o Excel
    filepath = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not filepath:
        CTkMessagebox(title="Importação cancelada", message="Nenhum arquivo foi selecionado.", icon="info")
        return

    try:
        # Lê o Excel
        df = pd.read_excel(filepath)

        # Verifica se as colunas necessárias existem
        if 'Nome' not in df.columns or 'Ultimo Backup' not in df.columns:
            CTkMessagebox(title="Erro", message="O arquivo deve conter as colunas 'Nome' e 'Ultimo Backup'.", icon="cancel")
            return

        atualizados = 0

        for index, row in df.iterrows():
            nome = str(row['Nome']).strip().lower()
            ultimo_bkp = str(row['Ultimo Backup']).strip()

            conta = session.query(Conta).filter_by(nome=nome).first()

            if conta:
                conta.ultimo_bkp = ultimo_bkp
                atualizados += 1

        session.commit()
        CTkMessagebox(title="Importação concluída", message=f"{atualizados} contas atualizadas com sucesso.", icon="check")

    except Exception as e:
        CTkMessagebox(title="Erro", message=f"Falha ao importar: {e}", icon="cancel")

    